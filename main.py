"""
论文评审系统 - 主程序入口
"""

import argparse
import os
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.upper() != "UTF-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

import yaml
from dotenv import load_dotenv

from src.paper_parser import PaperParser
from src.completeness_checker import CompletenessChecker
from src.image_analyzer import ImageAnalyzer
from src.ai_evaluator import AIEvaluator
from src.aigc_detector import AIGCDetector
from src.plagiarism_checker import PlagiarismChecker
from src.report_generator import ReportGenerator
from src.standards_generator import StandardsGenerator

try:
    from openpyxl import load_workbook
except ImportError:
    pass  # --post-normalize 模式下会报更明确的错误


def load_env():
    """加载 .env 文件并返回配置字典"""
    load_dotenv()

    providers = []
    for key_suffix in ["", "_2", "_3", "_4", "_5"]:
        base_url = os.getenv(f"API_BASE_URL{key_suffix}")
        api_key = os.getenv(f"API_KEY{key_suffix}")
        model = os.getenv(f"API_MODEL{key_suffix}")
        if base_url and api_key and model:
            providers.append({
                "base_url": base_url,
                "api_key": api_key,
                "model": model,
            })

    return {
        "providers": providers,
        "requirements_doc": os.getenv("REQUIREMENTS_DOC", ""),
        "output_dir": os.getenv("OUTPUT_DIR", "./outputs"),
        "papers_dir": os.getenv("PAPERS_DIR", ""),
        # 评分范围，可在 .env 覆盖 settings.yaml
        "score_internal_min": int(os.getenv("SCORE_INTERNAL_MIN", "")) if os.getenv("SCORE_INTERNAL_MIN", "") else None,
        "score_internal_max": int(os.getenv("SCORE_INTERNAL_MAX", "")) if os.getenv("SCORE_INTERNAL_MAX", "") else None,
        "output_min": int(os.getenv("OUTPUT_MIN", "")) if os.getenv("OUTPUT_MIN", "") else None,
        "output_max": int(os.getenv("OUTPUT_MAX", "")) if os.getenv("OUTPUT_MAX", "") else None,
        "output_exponent": float(os.getenv("OUTPUT_EXPONENT", "")) if os.getenv("OUTPUT_EXPONENT", "") else None,
    }


def load_settings() -> dict:
    """从 config/settings.yaml 加载配置"""
    settings_path = Path("config/settings.yaml")
    if settings_path.exists():
        with open(settings_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    return {}


def get_score_range(settings: dict, args, env_config: dict) -> dict:
    """获取分数范围：内部 0-100（优先级：args > .env > settings.yaml > 默认值）
    注：归一化（范围映射、指数）在生成报告后通过 --post-normalize 独立处理。
    """
    sr = settings.get("score_range", {})
    out_r = settings.get("output_range", {})

    internal_min = (
        args.score_min if args.score_min is not None
        else env_config.get("score_internal_min") or sr.get("min", 0)
    )
    internal_max = (
        args.score_max if args.score_max is not None
        else env_config.get("score_internal_max") or sr.get("max", 100)
    )

    # 归一化默认恒等映射（不修改分数）；调整在 --post-normalize 中独立进行
    output_min = 0
    output_max = 100
    output_exponent = 1.0
    return {
        "internal_min": internal_min,
        "internal_max": internal_max,
        "output_min": output_min,
        "output_max": output_max,
        "output_exponent": output_exponent,
    }


def load_requirements_config(requirements_path: str) -> dict:
    with open(requirements_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def _run_post_normalize(args):
    """读取已生成的 Excel 报告，对最终得分做后处理归一化"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("✗ 需要 openpyxl 库来读取 Excel，请执行: pip install openpyxl")
        sys.exit(1)

    output_dir = Path(args.post_normalize if isinstance(args.post_normalize, str) else "./outputs")
    if not output_dir.exists():
        print(f"✗ 输出目录不存在: {output_dir}")
        sys.exit(1)

    # 寻找最新的 scores_summary.xlsx
    xlsx_files = sorted(output_dir.glob("scores_summary*.xlsx"))
    if not xlsx_files:
        print(f"✗ 未找到 scores_summary.xlsx 文件: {output_dir}")
        sys.exit(1)
    src_path = xlsx_files[-1]

    print(f"  读取报告: {src_path}")
    wb = load_workbook(src_path)
    ws = wb.active

    # 找到表头行
    headers = [cell.value for cell in ws[1]]
    try:
        score_col = headers.index("最终得分") + 1  # 1-based
    except ValueError:
        print(f"✗ 未找到 '最终得分' 列，表头: {headers}")
        sys.exit(1)

    # 收集原始分数
    raw_scores = []
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[score_col - 1]
        if cell.value is None or not isinstance(cell.value, (int, float)):
            continue
        raw_scores.append(cell.value)
        rows.append(row)

    if not raw_scores:
        print("✗ 未找到有效的分数数据")
        sys.exit(1)

    # 获取归一化参数
    output_min = args.output_min if args.output_min is not None else 60
    output_max = args.output_max if args.output_max is not None else 90
    output_exponent = args.output_exponent if args.output_exponent is not None else 1.5

    print(f"  原始分数: min={min(raw_scores):.0f} max={max(raw_scores):.0f} (共 {len(raw_scores)} 个)")
    print(f"  归一化: {output_min} ~ {output_max}, exponent={output_exponent}")
    print()

    # 对每个分数做归一化
    rg = ReportGenerator(str(output_dir), output_min=output_min, output_max=output_max,
                         output_exponent=output_exponent)
    for i, row in enumerate(rows):
        raw = row[score_col - 1].value
        normalized = rg._normalize_score(raw)
        row[score_col - 1].value = normalized

    # 添加归一化说明（备注列或统计信息区域）
    stats_row = len(rows) + 3
    ws.cell(row=stats_row, column=1, value=f"归一化参数: 原始→{output_min}~{output_max}, exponent={output_exponent}")
    ws.cell(row=stats_row + 1, column=1, value="(归一化前原始分数可在原始报告中查看)")

    # 保存为新文件
    stem = src_path.stem.replace("scores_summary", "scores_normalized")
    out_path = src_path.with_name(f"{stem}.xlsx")
    wb.save(out_path)
    print(f"  ✓ 归一化报告已保存: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="论文评审系统")
    parser.add_argument(
        "--papers", "-p",
        default=None,
        help="学生论文文件夹路径（包含.docx/.doc/.pdf文件），不填则使用 .env 中的 PAPERS_DIR"
    )
    parser.add_argument(
        "--config", "-c",
        default="config/requirements.yaml",
        help="评分标准配置文件路径"
    )
    parser.add_argument(
        "--requirements-doc", "-r",
        default=None,
        help="题目要求Word文档路径（覆盖.env中的配置）"
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="输出目录（覆盖.env中的配置）"
    )
    parser.add_argument(
        "--skip-ai",
        action="store_true",
        help="跳过AI评审（仅做格式检测和查重）"
    )
    parser.add_argument(
        "--plagiarism",
        action="store_true",
        help="启用查重检测（默认关闭）"
    )
    parser.add_argument(
        "--generate-standards", "-g",
        action="store_true",
        help="根据题目要求自动生成评分标准（保存到 requirements.yaml），然后退出"
    )
    parser.add_argument(
        "--force-standards",
        action="store_true",
        help="强制重新生成评分标准（覆盖已有配置）"
    )
    parser.add_argument(
        "--skip-standards",
        action="store_true",
        help="跳过自动生成评分标准，使用已有配置文件"
    )
    parser.add_argument(
        "--score-min",
        type=int,
        default=None,
        help="内部最低分数（默认 0）"
    )
    parser.add_argument(
        "--score-max",
        type=int,
        default=None,
        help="内部最高分数（默认 100，AI评审以这个范围为基准打分）"
    )
    parser.add_argument(
        "--post-normalize",
        nargs="?",
        const=True,
        default=False,
        metavar="OUTPUT_DIR",
        help="对已生成的报告做后处理归一化。可指定输出目录（默认 ./outputs），需配合 --output-min/--output-max/--output-exponent 使用"
    )
    parser.add_argument(
        "--output-min",
        type=int,
        default=None,
        help="后处理归一化：输出最低分（与 --post-normalize 配合使用）"
    )
    parser.add_argument(
        "--output-max",
        type=int,
        default=None,
        help="后处理归一化：输出最高分（与 --post-normalize 配合使用）"
    )
    parser.add_argument(
        "--output-exponent",
        type=float,
        default=None,
        help="后处理归一化：映射指数，1.0=线性 >1=高分更难获得（与 --post-normalize 配合使用）"
    )
    args = parser.parse_args()

    # ══════════════════════════════════════════════════
    # 后处理归一化模式
    # ══════════════════════════════════════════════════
    if args.post_normalize is not False:
        _run_post_normalize(args)
        return

    print("=" * 60)
    print("论文评审系统 v1.0")
    print("=" * 60)

    print("\n[1/7] 加载配置...")
    env_config = load_env()
    settings = load_settings()
    score_ranges = get_score_range(settings, args, env_config)
    score_min = score_ranges["internal_min"]
    score_max = score_ranges["internal_max"]
    output_min = score_ranges["output_min"]
    output_max = score_ranges["output_max"]
    print(f"  评分范围: {score_min} ~ {score_max}（内部原始分数）")
    print(f"  💡 归一化（范围映射/指数）在报告生成后通过 --post-normalize 独立处理")
    ai_temperature = settings.get("ai_evaluation", {}).get("temperature", 0.4)
    print(f"  AI评审temperature: {ai_temperature}")
    output_exponent = 1.0

    requirements_doc_path = args.requirements_doc or env_config["requirements_doc"]
    if not requirements_doc_path:
        print("  ✗ 未指定题目要求文档，请通过 --requirements-doc 参数或 .env 文件配置 REQUIREMENTS_DOC")
        sys.exit(1)

    print(f"  读取题目要求文档: {requirements_doc_path}")
    requirements_text = PaperParser.parse_requirements_doc(requirements_doc_path)
    print(f"  ✓ 题目要求已读取 ({len(requirements_text)}字符)")

    providers = env_config.get("providers", [])
    if not providers:
        print("  ✗ 未配置 API，请检查 .env 中的 API_BASE_URL / API_KEY / API_MODEL")
        sys.exit(1)
    print(f"  ✓ 已加载 {len(providers)} 个 API Provider: {', '.join(p['model'] for p in providers)}")

    from src.llm_client import LLMClient
    text_llm = LLMClient(providers)

    # 筛选视觉模态 Provider（模型名含 agnes/vision/multimodal）
    vision_providers = [
        p for p in providers
        if any(kw in p["model"].lower() for kw in ("agnes", "vision", "multimodal"))
    ]
    vision_llm = LLMClient(vision_providers) if vision_providers else None
    if vision_llm:
        print(f"  ✓ 视觉分析已启用: {', '.join(p['model'] for p in vision_providers)}")
    else:
        print("  - 未配置视觉模态 Provider，图片分析将使用启发式规则")

    config_path = Path(args.config)

    if args.generate_standards and not args.force_standards and config_path.exists():
        print("\n[生成评分标准] 正在分析题目要求...")
        generator = StandardsGenerator(text_llm)
        result = generator.generate_and_save(requirements_text, args.config)
        if result.success:
            print("\n  ✓ 评分标准生成完成！")
            print(f"  维度: {', '.join(d['name'] for d in result.dimensions)}")
            sec_names = [s['name'] if isinstance(s, dict) else s for s in result.sections]
            print(f"  章节: {', '.join(sec_names)}")
            print(f"  最少字数: {result.min_word_count}")
            print(f"\n  已保存至 {args.config}，请检查后重新运行评审。")
        else:
            print(f"\n  ✗ 生成失败: {result.error_message}")
        return

    if not args.skip_standards:
        if args.force_standards:
            print("\n[生成评分标准] 强制重新生成...")
        else:
            print("\n[生成评分标准] 根据题目要求自动生成...")

        generator = StandardsGenerator(text_llm)
        result = generator.generate_and_save(requirements_text, args.config)
        if result.success:
            print("  ✓ 评分标准已生成")
            if result.dimensions:
                print(f"  维度: {', '.join(d['name'] for d in result.dimensions)}")
                for d in result.dimensions:
                    print(f"    - {d['name']}（{d['weight']*100:.0f}%）：{d['description']}")
            sec_names = [s['name'] if isinstance(s, dict) else s for s in result.sections] if result.sections else []
            if sec_names:
                print(f"  必要章节: {', '.join(sec_names)}")
            print(f"  最少字数: {result.min_word_count}")
        else:
            print(f"  ✗ 生成失败: {result.error_message}")
            print("  将使用默认评分标准...")
            args.config = None
    else:
        print("  使用已有评分标准配置")

    if args.config and Path(args.config).exists():
        req_config = load_requirements_config(args.config)
    else:
        req_config = {
            "evaluation_criteria": "默认评分标准",
            "dimensions": [
                {"name": "内容质量", "weight": 0.30, "description": "内容深度和广度"},
                {"name": "技术能力", "weight": 0.25, "description": "技术选型和代码质量"},
                {"name": "文档规范", "weight": 0.20, "description": "报告结构和格式"},
                {"name": "创新性", "weight": 0.10, "description": "创新点"},
                {"name": "学术诚信", "weight": 0.15, "description": "原创性"},
            ],
        }

    dimensions = req_config.get("dimensions", [])
    if not dimensions:
        print("  ✗ 评分维度为空，请检查配置文件")
        sys.exit(1)
    print(f"  ✓ 评分维度: {', '.join(d['name'] for d in dimensions)}")

    papers_dir = args.papers or env_config["papers_dir"]
    if not papers_dir:
        print("  ✗ 未指定论文文件夹，请通过 --papers 参数或 .env 文件配置 PAPERS_DIR")
        sys.exit(1)

    output_dir = args.output or env_config["output_dir"]
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    print(f"  ✓ 输出目录: {output_path}")

    print("\n[2/7] 解析论文...")
    paper_parser = PaperParser()
    papers = paper_parser.parse_directory(papers_dir)

    if not papers:
        print("  ✗ 未找到任何文档（.docx/.doc/.pdf），请检查路径")
        sys.exit(1)

    # 按学号排序
    papers.sort(key=lambda p: p.student_name)
    print("\n  学生列表（按学号排序）：")
    print(f"  {'序号':>4}  {'学号-姓名':<28}  {'文件':<50}")
    print(f"  {'-'*4}  {'-'*28}  {'-'*50}")
    for i, p in enumerate(papers, 1):
        print(f"  {i:>4}  {p.student_name:<28}  {p.filename:<50}")
    print(f"\n  ✓ 共解析 {len(papers)} 篇论文")

    # 确认
    try:
        resp = input("\n  确认上述学生信息无误？继续评审请输入 Y 或直接回车：").strip()
        if resp and resp.upper() != "Y":
            print("  ✗ 已取消")
            sys.exit(0)
    except (EOFError, KeyboardInterrupt):
        print("\n  ✗ 已取消")
        sys.exit(0)

    print("\n[3/7] 图片分析...")
    paper_image_analysis: dict[str, list] = {}
    if vision_llm:
        image_analyzer = ImageAnalyzer(vision_llm)
        for paper in papers:
            if paper.images:
                print(f"  - {paper.student_name}: {len(paper.images)}张图片")
                analysis = image_analyzer.analyze_images(paper.images)
                paper_image_analysis[paper.student_name] = analysis
                aigc_count = sum(1 for a in analysis if a.is_likely_aigc and a.aigc_confidence > 0.5)
                if aigc_count:
                    print(f"    → {aigc_count}张疑似AIGC生成")
        print(f"  ✓ 共分析 {sum(len(v) for v in paper_image_analysis.values())} 张图片")
    else:
        print("  - 跳过（未配置视觉模态 Provider，后续使用启发式规则）")

    print("\n[4/7] 完整性检测...")
    completeness_rules = req_config.get("completeness", {})
    if not completeness_rules or not completeness_rules.get("sections"):
        generated_meta = req_config.get("_generated", {})
        old_sections = generated_meta.get("sections", [
            "项目概述", "功能实现", "技术栈", "项目结构",
            "代码实现", "项目演示截图", "遇到的问题与解决方案", "总结与展望"
        ])
        old_min_words = generated_meta.get("min_word_count", 2000)
        completeness_rules = {
            "sections_weight": 40,
            "sections": [{"name": s, "patterns": [s], "weight": 0} for s in old_sections],
            "word_count": {"min": old_min_words, "weight": 20},
            "figures": {"min": 3, "weight": 15},
            "format": {"min_paragraphs": 10, "max_long_line_ratio": 0.3, "weight": 15},
        }

    completeness_checker = CompletenessChecker(completeness_rules)
    completeness_results = []
    for paper in papers:
        result = completeness_checker.check(paper)
        completeness_results.append(result)
        status = "✓" if result.is_complete else "⚠"
        details = []
        if result.missing_sections:
            details.append(f"缺{len(result.missing_sections)}章")
        print(f"  {status} {paper.student_name}: {result.score:.0f}分{' - ' + ', '.join(details) if details else ''}")

    print("\n[5/7] AIGC检测...")
    aigc_config = settings.get("aigc_detection", {})
    aigc_detector = AIGCDetector({
        "threshold": aigc_config.get("threshold", 0.6),
        "fake_impl_weight": aigc_config.get("fake_impl_weight", 0.5),
        "ai_writing_weight": aigc_config.get("ai_writing_weight", 0.4),
        "image_analysis_weight": aigc_config.get("image_analysis_weight", 0.25),
        "ai_patterns": [],
    })
    aigc_results = []
    for paper in papers:
        result = aigc_detector.detect(paper, image_analysis=paper_image_analysis.get(paper.student_name))
        aigc_results.append(result)
        icons = {"高风险": "🔴", "中风险": "🟡", "低风险": "🟢", "正常": "✅"}
        print(f"  {icons.get(result.overall_risk, '❓')} {paper.student_name}: "
              f"{result.overall_risk} (AI概率{result.ai_probability:.0%})")

    plagiarism_results = []
    if args.plagiarism:
        print("\n[6/7] 查重检测...")
        plagiarism_checker = PlagiarismChecker(
            config={
                "suspect_threshold": 0.4,
                "high_threshold": 0.6,
                "min_match_length": 30,
                "char_ngram_size": 8,
                "word_ngram_size": 3,
                "minhash_hashes": 128,
                "ai_check": True,
            },
            llm_client=text_llm,
        )
        plagiarism_results = plagiarism_checker.check_all(papers)

        plag_summary = plagiarism_checker.get_plagiarism_summary(plagiarism_results)
        if plag_summary["suspected_plagiarism_count"] > 0:
            print(f"  ⚠ 发现 {plag_summary['suspected_plagiarism_count']} 对疑似抄袭:")
            for case in plag_summary["details"]:
                print(f"    - {case['student']} ↔ {case['similar_to']} ({case['similarity']:.0%})")
            for r in plagiarism_results:
                for pair in r.suspicious_pairs:
                    other = pair.student_b if pair.student_a == r.student_name else pair.student_a
                    if pair.severity == "高度疑似":
                        print(f"    🔴 {r.student_name} ↔ {other}: {pair.similarity:.0%}（{pair.severity}）")
                    elif pair.severity == "疑似":
                        print(f"    🟡 {r.student_name} ↔ {other}: {pair.similarity:.0%}（{pair.severity}）")
        else:
            print("  ✓ 未发现明显抄袭")
    else:
        print("\n[6/7] 跳过查重检测")

    evaluation_results = []
    if not args.skip_ai:
        print("\n[7/7] AI评审（这可能需要一些时间）...")
        ai_evaluator = AIEvaluator(text_llm, score_min=score_min, score_max=score_max,
                                     temperature=ai_temperature)
        evaluation_results = ai_evaluator.evaluate_batch(
            papers=papers,
            requirements=requirements_text,
            evaluation_criteria=req_config.get("evaluation_criteria", ""),
            dimensions=dimensions,
            image_analysis_map=paper_image_analysis,
        )
        print("  ✓ AI评审完成")
    else:
        print("\n[7/7] 跳过AI评审")

    print("\n" + "=" * 60)
    print("生成报告...")
    print("=" * 60)

    report_gen = ReportGenerator(
        output_dir,
        score_min=score_min,
        score_max=score_max,
        output_min=output_min,
        output_max=output_max,
        output_exponent=output_exponent,
    )

    excel_path = report_gen.generate_excel(
        papers=papers,
        completeness_results=completeness_results,
        evaluation_results=evaluation_results,
        aigc_results=aigc_results,
        plagiarism_results=plagiarism_results,
        dimensions=dimensions,
    )
    print(f"\n  📊 Excel汇总表: {excel_path}")

    md_path = report_gen.generate_markdown_report(
        papers=papers,
        completeness_results=completeness_results,
        evaluation_results=evaluation_results,
        aigc_results=aigc_results,
        plagiarism_results=plagiarism_results,
        dimensions=dimensions,
    )
    print(f"  📄 MD评审报告: {md_path}")

    print("\n" + "=" * 60)
    print("评审完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()

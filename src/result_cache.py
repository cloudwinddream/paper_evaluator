"""
结果缓存 / 已有报告加载模块
- 优先从 .result_cache/ (JSON) 加载完整结果
- 回退从 scores_summary.xlsx 加载已有学生列表及部分结果数据
"""

import json
import re
from pathlib import Path

from src.ai_evaluator import EvaluationResult
from src.completeness_checker import CompletenessResult
from src.aigc_detector import AIGCResult
from src.plagiarism_checker import PlagiarismResult, PlagiarismPair


def _pair_to_dict(pair: PlagiarismPair) -> dict:
    return {"student_a": pair.student_a, "student_b": pair.student_b,
            "similarity": pair.similarity, "severity": pair.severity, "details": pair.details}


def _dict_to_pair(d: dict) -> PlagiarismPair:
    return PlagiarismPair(student_a=d["student_a"], student_b=d["student_b"],
                          similarity=d["similarity"], severity=d["severity"], details=d.get("details", ""))


# ── JSON 缓存读写 ──

def save_results(output_dir: str, evaluation_results: list[EvaluationResult],
                 completeness_results: list[CompletenessResult],
                 aigc_results: list[AIGCResult],
                 plagiarism_results: list[PlagiarismResult]):
    cache_dir = Path(output_dir) / ".result_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)

    eval_data = [{"student_name": r.student_name, "total_score": r.total_score,
                   "dimension_scores": r.dimension_scores, "evaluation_basis": r.evaluation_basis,
                   "short_comment": r.short_comment, "core_problems": r.core_problems,
                   "raw_response": r.raw_response, "success": r.success, "error_message": r.error_message}
                 for r in evaluation_results]
    (cache_dir / "evaluation.json").write_text(json.dumps(eval_data, ensure_ascii=False), encoding="utf-8")

    comp_data = [{"student_name": r.student_name, "is_complete": r.is_complete, "score": r.score,
                   "missing_sections": r.missing_sections, "warnings": r.warnings,
                   "details": r.details, "dimension_scores": r.dimension_scores}
                 for r in completeness_results]
    (cache_dir / "completeness.json").write_text(json.dumps(comp_data, ensure_ascii=False), encoding="utf-8")

    aigc_data = [{"student_name": r.student_name, "ai_probability": r.ai_probability,
                   "is_suspicious": r.is_suspicious, "suspicious_segments": r.suspicious_segments,
                   "ai_pattern_count": r.ai_pattern_count, "format_issues": r.format_issues,
                   "image_issues": r.image_issues, "overall_risk": r.overall_risk}
                 for r in aigc_results]
    (cache_dir / "aigc.json").write_text(json.dumps(aigc_data, ensure_ascii=False), encoding="utf-8")

    plag_data = [{"student_name": r.student_name, "highest_similarity": r.highest_similarity,
                   "most_similar_student": r.most_similar_student,
                   "suspicious_pairs": [_pair_to_dict(p) for p in r.suspicious_pairs]}
                 for r in plagiarism_results]
    (cache_dir / "plagiarism.json").write_text(json.dumps(plag_data, ensure_ascii=False), encoding="utf-8")


def _load_json_cache(output_dir: str):
    """从 JSON 缓存加载完整结果"""
    cache_dir = Path(output_dir) / ".result_cache"
    if not cache_dir.exists():
        return None

    evaluation_results = []
    eval_path = cache_dir / "evaluation.json"
    if eval_path.exists():
        for d in json.loads(eval_path.read_text(encoding="utf-8")):
            evaluation_results.append(EvaluationResult(**d))

    completeness_results = []
    comp_path = cache_dir / "completeness.json"
    if comp_path.exists():
        for d in json.loads(comp_path.read_text(encoding="utf-8")):
            completeness_results.append(CompletenessResult(**d))

    aigc_results = []
    aigc_path = cache_dir / "aigc.json"
    if aigc_path.exists():
        for d in json.loads(aigc_path.read_text(encoding="utf-8")):
            aigc_results.append(AIGCResult(**d))

    plagiarism_results = []
    plag_path = cache_dir / "plagiarism.json"
    if plag_path.exists():
        for d in json.loads(plag_path.read_text(encoding="utf-8")):
            pairs = [_dict_to_pair(p) for p in d.pop("suspicious_pairs", [])]
            r = PlagiarismResult(**d)
            r.suspicious_pairs = pairs
            plagiarism_results.append(r)

    return evaluation_results, completeness_results, aigc_results, plagiarism_results


# ── 从已有 Excel 报告恢复部分结果 ──

def _parse_excel_results(output_dir: str, dimensions: list[dict]):
    """从 scores_summary.xlsx 读取已有学生名单及部分结果数据"""
    xlsx_path = _find_excel_report(output_dir)
    if not xlsx_path:
        return None

    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # 定位关键列
    col_map = {}
    for i, h in enumerate(headers, 1):
        col_map[h] = i

    name_col = col_map.get("学生姓名")
    comp_col = col_map.get("完整性得分")
    total_col = col_map.get("AI评审总分")
    risk_col = col_map.get("AIGC风险")
    deduction_col = col_map.get("AIGC扣分")
    plag_col = col_map.get("查重最高相似度")
    comment_col = col_map.get("简短评语")
    detail_col = col_map.get("详细评语")

    if not name_col:
        return None

    # 维度列：按位置匹配（列名可能随配置变化）
    dim_start = comp_col + 1 if comp_col else 6  # 完整性得分之后的第一列
    dim_end = total_col - 1 if total_col else (dim_start + len(dimensions) - 1)

    # 解析每一行
    eval_results = []
    comp_results = []
    aigc_results = []
    plag_results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_col - 1] if name_col and name_col <= len(row) else None
        if not name:
            continue

        # CompletenessResult
        comp_score = row[comp_col - 1] if comp_col and comp_col <= len(row) else 0
        if isinstance(comp_score, str):
            comp_score = 0
        comp_results.append(CompletenessResult(
            student_name=str(name), score=float(comp_score or 0),
            is_complete=True,
        ))

        # EvaluationResult (按位置读取维度分数)
        dim_scores = {}
        for idx, dn in enumerate([d["name"] for d in dimensions]):
            col = dim_start + idx
            if col <= len(row):
                v = row[col - 1]
                if isinstance(v, (int, float)):
                    dim_scores[dn] = int(v)
        total = row[total_col - 1] if total_col and total_col <= len(row) else 0
        short_comment = row[comment_col - 1] if comment_col and comment_col <= len(row) else ""
        detailed = str(row[detail_col - 1] or "") if detail_col and detail_col <= len(row) else ""

        eval_results.append(EvaluationResult(
            student_name=str(name), total_score=float(total or 0),
            dimension_scores=dim_scores,
            short_comment=str(short_comment or ""),
            evaluation_basis=str(detailed or ""),
            success=True,
        ))

        # AIGCResult
        risk = str(row[risk_col - 1]) if risk_col and risk_col <= len(row) else "正常"
        deduction_str = str(row[deduction_col - 1]) if deduction_col and deduction_col <= len(row) else "0%"
        deduction_match = re.search(r"(\d+)", deduction_str)
        deduction = int(deduction_match.group(1)) if deduction_match else 0
        ai_prob = min(deduction / 80.0, 1.0) if deduction > 0 else 0.0
        is_suspicious = risk not in ("正常", "低风险")
        aigc_results.append(AIGCResult(
            student_name=str(name), ai_probability=ai_prob,
            is_suspicious=is_suspicious, overall_risk=risk,
        ))

        # PlagiarismResult
        plag_sim_str = str(row[plag_col - 1]) if plag_col and plag_col <= len(row) else "0%"
        sim_match = re.search(r"(\d+)", plag_sim_str)
        plag_sim = int(sim_match.group(1)) / 100.0 if sim_match else 0.0
        plag_results.append(PlagiarismResult(
            student_name=str(name), highest_similarity=plag_sim,
        ))

    return eval_results, comp_results, aigc_results, plag_results


def _find_excel_report(output_dir: str):
    """查找输出目录中最新的 scores_summary.xlsx"""
    out = Path(output_dir)
    candidates = sorted(out.glob("scores_summary*.xlsx"))
    return candidates[-1] if candidates else None


# ── 统一入口 ──

def load_existing_results(output_dir: str, dimensions: list[dict]):
    """加载已有评审结果，优先从 JSON 缓存，回退到 Excel"""
    cached = _load_json_cache(output_dir)
    if cached is not None:
        print("  [OK] found result cache")
        return cached

    excel = _parse_excel_results(output_dir, dimensions)
    if excel is not None:
        print("  [OK] recovered from Excel report")
        return excel

    return [], [], [], []


def get_evaluated_names(output_dir: str, dimensions: list[dict]):
    """仅获取已有评审的学生姓名列表（不构造结果对象）"""
    cached = _load_json_cache(output_dir)
    if cached is not None:
        return {r.student_name for r in cached[0]}

    # 快速：直接从 Excel 读姓名列
    xlsx_path = _find_excel_report(output_dir)
    if not xlsx_path:
        return set()

    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        try:
            name_idx = headers.index("学生姓名") + 1  # 1-based
        except ValueError:
            return set()
        names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[name_idx - 1] is not None:
                names.add(str(row[name_idx - 1]))
        return names
    except Exception:
        return set()
